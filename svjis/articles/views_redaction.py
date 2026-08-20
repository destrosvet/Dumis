from . import utils, models, forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.models import Group, User
from django.core.paginator import Paginator, InvalidPage
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone, dateformat
from django.utils.timezone import make_aware
from django.utils.translation import gettext_lazy as _
from django.utils.translation import gettext as gt
from django.views.decorators.http import require_GET, require_POST
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from .model_utils import PICTURE_ICONS, resize_uploaded_image
from .user_agent import get_browser, get_os
from .permissions import (
    svjis_view_redaction_menu,
    svjis_edit_article,
    svjis_edit_survey,
    svjis_edit_article_menu,
    svjis_edit_article_news,
    svjis_edit_useful_link,
)
from urllib.parse import quote
from django.utils.html import escape
from django.utils.safestring import mark_safe
from .templatetags.article_filters import highlight, yes_no


def get_side_menu(active_item, user):
    side_menu = [
        {
            'perms': svjis_edit_article,
            'description': _("Articles"),
            'link': reverse(redaction_article_view),
            'active': True if active_item == 'article' else False,
        },
        {
            'perms': svjis_edit_article_news,
            'description': _("News"),
            'link': reverse(redaction_news_view),
            'active': True if active_item == 'news' else False,
        },
        {
            'perms': svjis_edit_useful_link,
            'description': _("Useful Links"),
            'link': reverse(redaction_useful_link_view),
            'active': True if active_item == 'links' else False,
        },
        {
            'perms': svjis_edit_survey,
            'description': _("Surveys"),
            'link': reverse(redaction_survey_view),
            'active': True if active_item == 'surveys' else False,
        },
        {
            'perms': svjis_edit_article_menu,
            'description': _("Menu"),
            'link': reverse(redaction_menu_view),
            'active': True if active_item == 'menu' else False,
        },
        {
            'perms': svjis_view_redaction_menu,
            'description': _("Analytics"),
            'link': reverse(redaction_analytics_view),
            'active': True if active_item == 'analytics' else False,
        },
    ]
    return [x for x in side_menu if x['perms'] is None or user.has_perm(x['perms'])]


# Redaction - Article Menu
def get_article_menu():
    result = []
    menu_items = models.ArticleMenu.objects.all()
    level = 0
    for i in menu_items:
        if i.parent is None:
            node = {'item': i, 'level': level, 'articles': i.articles.count()}
            result.append(node)
            result.extend(get_article_submenu(i, menu_items, level + 1))
    return result


def get_article_submenu(parent, menu_items, level):
    result = []
    for i in menu_items:
        if i.parent == parent:
            node = {'item': i, 'level': level, 'articles': i.articles.count()}
            result.append(node)
            result.extend(get_article_submenu(i, menu_items, level + 1))
    return result


@permission_required(svjis_edit_article_menu)
@require_GET
def redaction_menu_view(request):
    columns = [
        {'key': 'name', 'label': _('Name')},
        {'key': 'parent', 'label': _('Parent')},
        {'key': 'level', 'label': _('Level')},
        {'key': 'hide', 'label': _('Hide')},
        {'key': 'articles', 'label': _('Articles')},
    ]
    rows = []
    for obj in get_article_menu():
        actions = [
            {'url': reverse('redaction_menu_edit', args=[obj['item'].pk]), 'icon': 'edit', 'label': _('Edit')},
        ]
        if obj['articles'] == 0:
            actions.append(
                {
                    'url': reverse('redaction_menu_delete', args=[obj['item'].pk]),
                    'icon': 'delete',
                    'label': _('Delete'),
                    'danger': True,
                    'confirm': f"{_('Do you want to delete')} {obj['item'].description} ?",
                }
            )
        name = escape(obj['item'].description)
        if obj['level'] > 1:
            name = f'<span style="background-color:#f1948a;">{name}</span>'
        rows.append(
            {
                'url': reverse('redaction_menu_edit', args=[obj['item'].pk]),
                'cells': {
                    'name': mark_safe('&nbsp;' * (3 * obj['level']) + name),
                    'parent': obj['item'].parent.description if obj['item'].parent else '',
                    'level': obj['level'],
                    'hide': yes_no(obj['item'].hide),
                    'articles': obj['articles'],
                },
                'actions': actions,
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['aside_menu_items'] = get_side_menu('menu', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    ctx['object_list'] = get_article_menu()
    ctx['columns'] = columns
    ctx['rows'] = rows
    ctx['table_id'] = 'menu'
    ctx['desc_id'] = 'menu'
    ctx['list_search'] = True
    return render(request, "redaction_menu.html", ctx)


@permission_required(svjis_edit_article_menu)
@require_GET
def redaction_menu_edit_view(request, pk):
    if pk:
        am = get_object_or_404(models.ArticleMenu, pk=pk)
        form = forms.ArticleMenuForm(instance=am)
    else:
        form = forms.ArticleMenuForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('menu', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_menu_edit.html", ctx)


@permission_required(svjis_edit_article_menu)
@require_POST
def redaction_menu_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.ArticleMenuForm(request.POST)
    else:
        instance = get_object_or_404(models.ArticleMenu, pk=pk)
        form = forms.ArticleMenuForm(request.POST, instance=instance)

    if form.is_valid():
        form.save()
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")

    return redirect(redaction_menu_view)


@permission_required(svjis_edit_article_menu)
@require_GET
def redaction_menu_delete_view(request, pk):
    obj = get_object_or_404(models.ArticleMenu, pk=pk)
    obj.delete()
    return redirect(redaction_menu_view)


# Redaction - Article
@permission_required(svjis_edit_article)
@require_GET
def redaction_article_view(request):
    article_list = models.Article.objects.select_related('author', 'menu')

    # Search
    search = request.GET.get('search')
    if search is not None and len(search) < 3:
        messages.error(request, _("Search: Keyword '{}' is too short. Type at least 3 characters.").format(search))
        search = None
    if search is not None and len(search) > 100:
        messages.error(request, _("Search: Keyword is too long. Type maximum of 100 characters."))
        search = None
    if search is not None:
        article_list = article_list.filter(
            Q(header__icontains=search) | Q(perex__icontains=search) | Q(body__icontains=search)
        )
        header = _("Search results") + f": {search}"
    else:
        search = ''
        header = _("Article")

    # Paginator
    is_paginated = len(article_list) > getattr(settings, 'SVJIS_ARTICLE_PAGE_SIZE', 10)
    page = request.GET.get('page', 1)
    paginator = Paginator(article_list, per_page=getattr(settings, 'SVJIS_ARTICLE_PAGE_SIZE', 10))
    page_obj = paginator.get_page(page)
    try:
        article_list = paginator.page(page)
    except InvalidPage:
        article_list = paginator.page(paginator.num_pages)
    page_parameter = '' if search == '' else f"search={search}"

    columns = [
        {'key': 'header', 'label': _('Article')},
        {'key': 'menu', 'label': _('Menu'), 'hide_on_mobile': True},
        {'key': 'author', 'label': _('Author'), 'hide_on_mobile': True},
        {'key': 'date', 'label': _('Date'), 'hide_on_mobile': True},
        {'key': 'published', 'label': _('Published'), 'hide_on_mobile': True},
    ]
    rows = []
    for a in article_list:
        url = reverse('article', args=[a.slug])
        if search:
            url += '?search=' + quote(search)
        rows.append(
            {
                'url': url,
                'cells': {
                    'header': highlight(a.header, search),
                    'menu': a.menu.description,
                    'author': f"{a.author.first_name} {a.author.last_name}",
                    'date': dateformat.format(a.published_date, 'd.m.Y') if a.published_date else '',
                    'published': yes_no(a.published),
                },
                'actions': [
                    {
                        'url': reverse('redaction_article_edit', args=[a.pk]),
                        'icon': 'edit',
                        'label': _('Edit'),
                    },
                    {
                        'url': reverse('redaction_article_notifications', args=[a.pk]),
                        'icon': 'email',
                        'label': _('Send notifications'),
                    },
                    {
                        'url': reverse('redaction_article_hide', args=[a.pk]),
                        'icon': 'view',
                        'label': _('Hide'),
                        'confirm': f"{_('Do you want to hide')} {a.header} ?",
                    },
                    {
                        'url': reverse('redaction_article_delete', args=[a.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete')} {a.header} ?",
                    },
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['is_paginated'] = is_paginated
    ctx['page_obj'] = page_obj
    ctx['page_parameter'] = page_parameter
    ctx['search_endpoint'] = reverse(redaction_article_view)
    ctx['search'] = search
    ctx['header'] = header
    ctx['aside_menu_items'] = get_side_menu('article', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    ctx['object_list'] = article_list
    ctx['columns'] = columns
    ctx['rows'] = rows
    ctx['table_id'] = 'articles'
    ctx['desc_id'] = 'articles'
    ctx['list_search'] = False
    return render(request, "redaction_article.html", ctx)


@permission_required(svjis_edit_article)
@require_GET
def redaction_article_edit_view(request, pk):
    if pk:
        a = get_object_or_404(models.Article, pk=pk)
        form = forms.ArticleForm(instance=a)
    else:
        form = forms.ArticleForm

    group_list = []
    for g in Group.objects.all().order_by('name'):
        item = {'name': g.name, 'checked': g in form.instance.visible_for_group.all() if pk else False}
        group_list.append(item)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['form'] = form
    ctx['group_list'] = group_list
    ctx['asset_form'] = forms.ArticleAssetForm
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('article', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_article_edit.html", ctx)


@permission_required(svjis_edit_article)
@require_POST
def redaction_article_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.ArticleForm(request.POST, request.FILES)
    else:
        instance = get_object_or_404(models.Article, pk=pk)
        form = forms.ArticleForm(request.POST, request.FILES, instance=instance)

    if form.is_valid():
        obj = form.save(commit=False)
        if not pk:
            obj.author = request.user
        if 'cover_image' in request.FILES:
            try:
                resized = resize_uploaded_image(obj.cover_image)
            except OSError:
                pass
            else:
                obj.cover_image.save(resized.name, resized, save=False)
        obj.save()
        pk = obj.pk

        # Set groups
        selected_group_names = set(request.POST.getlist('visible_for_groups'))
        current_group_names = set(obj.visible_for_group.values_list('name', flat=True))
        for g in Group.objects.all():
            if g.name in selected_group_names and g.name not in current_group_names:
                obj.visible_for_group.add(g)
            if g.name not in selected_group_names and g.name in current_group_names:
                obj.visible_for_group.remove(g)

        # Set watching users
        obj.watching_users.add(request.user)

        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")

    return redirect(reverse('redaction_article_edit', kwargs={'pk': pk}))


@permission_required(svjis_edit_article)
@require_GET
def redaction_article_hide_view(request, pk):
    article = get_object_or_404(models.Article, pk=pk)
    article.published = False
    article.save()
    return redirect(redaction_article_view)


@permission_required(svjis_edit_article)
@require_GET
def redaction_article_delete_view(request, pk):
    article = get_object_or_404(models.Article, pk=pk)
    article.delete()
    return redirect(redaction_article_view)


def get_users_for_notification(article):
    users = []
    if article.published:
        users = User.objects.filter(is_active=True).exclude(email='').distinct().order_by('last_name')
        if not article.visible_for_all:
            groups = article.visible_for_group.all()
            q = Q(groups__in=groups)
            users = users.filter(q)
    return users


@permission_required(svjis_edit_article)
@require_GET
def redaction_article_notifications_view(request, pk):
    article = get_object_or_404(models.Article, pk=pk)
    users = get_users_for_notification(article)

    ctx = utils.get_context()
    ctx['article'] = article
    ctx['object_list'] = users
    ctx['aside_menu_name'] = _("Redaction")
    ctx['aside_menu_items'] = get_side_menu('article', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_article_notifications.html", ctx)


@permission_required(svjis_edit_article)
@require_POST
def redaction_article_notifications_send_view(request):
    pk = int(request.POST['pk'])
    article = get_object_or_404(models.Article, pk=pk)
    users = get_users_for_notification(article)

    recipients = [u for u in users if request.POST.get(f"u_{u.pk}", False) == 'on']
    utils.send_article_notification(recipients, f"{request.scheme}://{request.get_host()}", article)
    i = len(recipients)

    ctx = utils.get_context()
    ctx['article'] = article
    ctx['num_of_recipients'] = i
    ctx['aside_menu_name'] = _("Redaction")
    ctx['aside_menu_items'] = get_side_menu('article', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_article_notifications_sent.html", ctx)


# Redaction - ArticleAsset
@permission_required(svjis_edit_article)
@require_POST
def redaction_article_asset_save_view(request):
    article_pk = int(request.POST.get('article_pk'))
    article = get_object_or_404(models.Article, pk=article_pk)
    form = forms.ArticleAssetForm(request.POST, request.FILES)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.article = article
        obj.save()
    else:
        for error in form.errors:
            messages.error(request, error)
    return redirect(reverse('redaction_article_edit', kwargs={'pk': article_pk}) + '#assets')


@permission_required(svjis_edit_article)
@require_POST
def redaction_article_image_upload_view(request):
    """Endpoint for images inserted via the HugeRTE editor in the article body.

    Stores the image as an ArticleAsset and returns its URL as JSON:
    {"location": "/media/..."}.
    """
    try:
        article_pk = int(request.POST.get('article_pk'))
    except (TypeError, ValueError):
        return JsonResponse({'error': gt('Article not found')}, status=400)
    article = models.Article.objects.filter(pk=article_pk).first()
    if article is None:
        return JsonResponse({'error': gt('Article not found')}, status=400)

    file = request.FILES.get('file')
    if file is None:
        return JsonResponse({'error': gt('No file was submitted')}, status=400)

    extension = os.path.splitext(file.name)[1][1:].lower()
    if extension not in PICTURE_ICONS:
        return JsonResponse({'error': gt('Unsupported image type')}, status=400)

    try:
        file = resize_uploaded_image(file)
    except OSError:
        return JsonResponse({'error': gt('Unsupported image type')}, status=400)

    asset = models.ArticleAsset.objects.create(article=article, description=file.name[:100], file=file)
    return JsonResponse({'location': f'/media/{asset.file}'})


@permission_required(svjis_edit_article)
@require_GET
def redaction_article_asset_delete_view(request, pk):
    obj = get_object_or_404(models.ArticleAsset, pk=pk)
    article_pk = obj.article.pk
    obj.delete()
    return redirect(reverse('redaction_article_edit', kwargs={'pk': article_pk}) + '#assets')


# Redaction - MiniNews
@permission_required(svjis_edit_article_news)
@require_GET
def redaction_news_view(request):
    news_list = models.News.objects.select_related('author')
    header = _("News")

    # Paginator
    is_paginated = len(news_list) > getattr(settings, 'SVJIS_NEWS_PAGE_SIZE', 10)
    page = request.GET.get('page', 1)
    paginator = Paginator(news_list, per_page=getattr(settings, 'SVJIS_NEWS_PAGE_SIZE', 10))
    page_obj = paginator.get_page(page)
    try:
        news_list = paginator.page(page)
    except InvalidPage:
        news_list = paginator.page(paginator.num_pages)

    columns = [
        {'key': 'date', 'label': _('Date'), 'hide_on_mobile': True},
        {'key': 'author', 'label': _('Author'), 'hide_on_mobile': True},
        {'key': 'published', 'label': _('Published')},
        {'key': 'body', 'label': _('Body')},
    ]
    rows = []
    for n in news_list:
        rows.append(
            {
                'url': reverse('redaction_news_edit', args=[n.pk]),
                'cells': {
                    'date': dateformat.format(n.created_date, 'd.m.Y H:i'),
                    'author': f"{n.author.first_name} {n.author.last_name}",
                    'published': yes_no(n.published),
                    'body': mark_safe(n.body),
                },
                'actions': [
                    {'url': reverse('redaction_news_edit', args=[n.pk]), 'icon': 'edit', 'label': _('Edit')},
                    {
                        'url': reverse('redaction_news_delete', args=[n.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete this news')}?",
                    },
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['is_paginated'] = is_paginated
    ctx['page_obj'] = page_obj
    ctx['header'] = header
    ctx['aside_menu_items'] = get_side_menu('news', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    ctx['object_list'] = news_list
    ctx['columns'] = columns
    ctx['rows'] = rows
    ctx['table_id'] = 'news'
    ctx['desc_id'] = 'news'
    ctx['list_search'] = True
    return render(request, "redaction_news.html", ctx)


@permission_required(svjis_edit_article_news)
@require_GET
def redaction_news_edit_view(request, pk):
    if pk:
        a = get_object_or_404(models.News, pk=pk)
        form = forms.NewsForm(instance=a)
    else:
        form = forms.NewsForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('news', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_news_edit.html", ctx)


@permission_required(svjis_edit_article_news)
@require_POST
def redaction_news_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.NewsForm(request.POST)
    else:
        instance = get_object_or_404(models.News, pk=pk)
        form = forms.NewsForm(request.POST, instance=instance)

    if form.is_valid():
        obj = form.save(commit=False)
        if not pk:
            obj.author = request.user
        obj.save()
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, error)
    return redirect(redaction_news_view)


@permission_required(svjis_edit_article_news)
@require_GET
def redaction_news_delete_view(request, pk):
    obj = get_object_or_404(models.News, pk=pk)
    obj.delete()
    return redirect(redaction_news_view)


# Redaction - UsefulLink
@permission_required(svjis_edit_useful_link)
@require_GET
def redaction_useful_link_view(request):
    useful_link_list = models.UsefulLink.objects.all()
    header = _("Useful Links")

    # Paginator
    is_paginated = len(useful_link_list) > getattr(settings, 'SVJIS_USEFUL_LIST_PAGE_SIZE', 10)
    page = request.GET.get('page', 1)
    paginator = Paginator(useful_link_list, per_page=getattr(settings, 'SVJIS_USEFUL_LIST_PAGE_SIZE', 10))
    page_obj = paginator.get_page(page)
    try:
        useful_link_list = paginator.page(page)
    except InvalidPage:
        useful_link_list = paginator.page(paginator.num_pages)

    columns = [
        {'key': 'header', 'label': _('Header')},
        {'key': 'order', 'label': _('Order')},
        {'key': 'published', 'label': _('Published')},
    ]
    rows = []
    for l in useful_link_list:
        rows.append(
            {
                'url': reverse('redaction_useful_link_edit', args=[l.pk]),
                'cells': {
                    'header': mark_safe(
                        f'<a href="{escape(l.link)}" onclick="event.stopPropagation()">{escape(l.header)}</a>'
                    ),
                    'order': l.order,
                    'published': yes_no(l.published),
                },
                'actions': [
                    {
                        'url': reverse('redaction_useful_link_edit', args=[l.pk]),
                        'icon': 'edit',
                        'label': _('Edit'),
                    },
                    {
                        'url': reverse('redaction_useful_link_delete', args=[l.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete this useful link')}?",
                    },
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['is_paginated'] = is_paginated
    ctx['page_obj'] = page_obj
    ctx['header'] = header
    ctx['aside_menu_items'] = get_side_menu('links', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    ctx['object_list'] = useful_link_list
    ctx['columns'] = columns
    ctx['rows'] = rows
    ctx['table_id'] = 'links'
    ctx['desc_id'] = 'links'
    ctx['list_search'] = True
    return render(request, "redaction_useful_link.html", ctx)


@permission_required(svjis_edit_useful_link)
@require_GET
def redaction_useful_link_edit_view(request, pk):
    if pk:
        a = get_object_or_404(models.UsefulLink, pk=pk)
        form = forms.UsefulLinkForm(instance=a)
    else:
        form = forms.UsefulLinkForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('links', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_useful_link_edit.html", ctx)


@permission_required(svjis_edit_useful_link)
@require_POST
def redaction_useful_link_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.UsefulLinkForm(request.POST)
    else:
        instance = get_object_or_404(models.UsefulLink, pk=pk)
        form = forms.UsefulLinkForm(request.POST, instance=instance)

    if form.is_valid():
        form.save()
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, error)
    return redirect(redaction_useful_link_view)


@permission_required(svjis_edit_useful_link)
@require_GET
def redaction_useful_link_delete_view(request, pk):
    obj = get_object_or_404(models.UsefulLink, pk=pk)
    obj.delete()
    return redirect(redaction_useful_link_view)


# Redaction - Surveys
@permission_required(svjis_edit_survey)
@require_GET
def redaction_survey_view(request):
    survey_list = models.Survey.objects.select_related('author')
    header = _("Surveys")

    # Paginator
    is_paginated = len(survey_list) > getattr(settings, 'SVJIS_SURVEY_PAGE_SIZE', 10)
    page = request.GET.get('page', 1)
    paginator = Paginator(survey_list, per_page=getattr(settings, 'SVJIS_SURVEY_PAGE_SIZE', 10))
    page_obj = paginator.get_page(page)
    try:
        survey_list = paginator.page(page)
    except InvalidPage:
        survey_list = paginator.page(paginator.num_pages)

    columns = [
        {'key': 'description', 'label': _('Description')},
        {'key': 'starting_date', 'label': _('Starting date'), 'hide_on_mobile': True},
        {'key': 'ending_date', 'label': _('Ending date'), 'hide_on_mobile': True},
        {'key': 'author', 'label': _('Author'), 'hide_on_mobile': True},
        {'key': 'published', 'label': _('Published')},
    ]
    rows = []
    for s in survey_list:
        rows.append(
            {
                'url': reverse('redaction_survey_edit', args=[s.pk]),
                'cells': {
                    'description': mark_safe(s.description),
                    'starting_date': dateformat.format(s.starting_date, 'd.m.Y'),
                    'ending_date': dateformat.format(s.ending_date, 'd.m.Y'),
                    'author': f"{s.author.first_name} {s.author.last_name}",
                    'published': yes_no(s.published),
                },
                'actions': [
                    {
                        'url': reverse('redaction_survey_edit', args=[s.pk]),
                        'icon': 'edit',
                        'label': _('Edit'),
                    },
                    {
                        'url': reverse('redaction_survey_results', args=[s.pk]),
                        'icon': 'chart',
                        'label': _('Results'),
                    },
                    {
                        'url': reverse('redaction_survey_results_export_to_excel', args=[s.pk]),
                        'icon': 'excel',
                        'label': _('Export results to Excel'),
                    },
                    {
                        'url': reverse('redaction_survey_delete', args=[s.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete this survey')} ?",
                    },
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['is_paginated'] = is_paginated
    ctx['page_obj'] = page_obj
    ctx['header'] = header
    ctx['aside_menu_items'] = get_side_menu('surveys', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    ctx['object_list'] = survey_list
    ctx['columns'] = columns
    ctx['rows'] = rows
    ctx['table_id'] = 'survey'
    ctx['desc_id'] = 'survey'
    ctx['list_search'] = True
    return render(request, "redaction_survey.html", ctx)


@permission_required(svjis_edit_survey)
@require_GET
def redaction_survey_edit_view(request, pk):
    if pk:
        a = get_object_or_404(models.Survey, pk=pk)
        form = forms.SurveyForm(instance=a)
        n = a.options.count()
    else:
        form = forms.SurveyForm
        n = 0

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['new_option_no'] = n + 1
    ctx['aside_menu_items'] = get_side_menu('surveys', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_survey_edit.html", ctx)


@permission_required(svjis_edit_survey)
@require_POST
def redaction_survey_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.SurveyForm(request.POST)
    else:
        instance = get_object_or_404(models.Survey, pk=pk)
        form = forms.SurveyForm(request.POST, instance=instance)

    if form.is_valid():
        obj = form.save(commit=False)
        if not pk:
            obj.author = request.user
        obj.save()

        # Options
        i = 1
        while f'oid_{i}' in request.POST:
            o_pk = int(request.POST[f'oid_{i}'])
            o_description = request.POST.get(f'o_{i}', '')
            if o_pk != 0:
                o_i = get_object_or_404(models.SurveyOption, pk=o_pk)
                o_i.description = o_description
                o_i.save()
            else:
                if o_description != '':
                    models.SurveyOption.objects.create(survey=obj, description=o_description)
            i += 1

        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, error)
    return redirect(redaction_survey_view)


@permission_required(svjis_edit_survey)
@require_GET
def redaction_survey_delete_view(request, pk):
    obj = get_object_or_404(models.Survey, pk=pk)
    obj.delete()
    return redirect(redaction_survey_view)


@permission_required(svjis_edit_survey)
@require_GET
def redaction_survey_option_delete_view(request, pk):
    obj = get_object_or_404(models.SurveyOption, pk=pk)
    survey_pk = obj.survey.pk
    obj.delete()
    return redirect(redaction_survey_edit_view, pk=survey_pk)


@permission_required(svjis_edit_survey)
@require_GET
def redaction_survey_results_view(request, pk):
    header = _("Survey")
    survey = get_object_or_404(models.Survey, pk=pk)
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['obj'] = survey
    ctx['header'] = header
    ctx['aside_menu_items'] = get_side_menu('surveys', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_survey_results.html", ctx)


@permission_required(svjis_edit_survey)
@require_GET
def redaction_survey_results_export_to_excel_view(request, pk):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Survey_Results.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = gt("Survey")

    survey = get_object_or_404(models.Survey, pk=pk)

    # Survey
    ws['A1'] = gt("Survey")
    ws['A1'].font = Font(bold=True)
    ws['A2'] = survey.description

    # Result
    ws['A4'] = gt("Result")
    ws['A4'].font = Font(bold=True)

    headers = [gt("Description"), gt("Votes"), gt("Votes") + ' %']
    ws.append(headers)

    header_st = utils.get_worksheet_header_style()
    for rows in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=len(headers)):
        for cell in rows:
            cell.style = header_st

    for o in survey.options:
        ws.append([o.description, o.total, round(o.pct, 2)])

    # Log
    last = len(ws['A']) + 1
    ws['A' + str(last + 1)] = gt("Voting log")
    ws['A' + str(last + 1)].font = Font(bold=True)

    headers = [gt("Time"), gt("User"), gt("Option")]
    ws.append(headers)

    last = len(ws['A'])
    for rows in ws.iter_rows(min_row=last, max_row=last, min_col=1, max_col=len(headers)):
        for cell in rows:
            cell.style = header_st

    for a in survey.answers:
        time = dateformat.format(timezone.localtime(a.time), "d.m.Y H:i")
        ws.append([time, f'{a.user.first_name} {a.user.last_name}', a.option.description])

    # Save the workbook to the HttpResponse
    utils.adjust_worksheet_columns_width(ws, 50)
    wb.save(response)
    return response


# Redaction - Analytics
@permission_required(svjis_view_redaction_menu)
@require_GET
def redaction_analytics_view(request):
    header = _("Analytics")
    history_in_days = getattr(settings, 'SVJIS_TOP_ARTICLES_HISTORY_IN_DAYS', 365)
    time_span_info = _("Data for last {} days.").format(history_in_days)
    top_history_from = make_aware(datetime.now() - timedelta(days=history_in_days))

    data = models.ArticleLog.objects.filter(entry_time__gte=top_history_from).exclude(user_agent='')
    scope = request.GET.get('scope', 'all')
    if scope == 'logged':
        data = data.exclude(user__isnull=True)
    data = data.values('user_agent').annotate(total=Count('*'))

    human_ua_table = []
    bot_ua_table = []
    browser_chart = {}
    os_chart = {}
    platform_chart = {}
    bot_chart = {}
    for d in data:
        browser = get_browser(d["user_agent"])["browser"]
        osystem_dict = get_os(d["user_agent"])
        osystem = osystem_dict["os"]
        platform = osystem_dict["platform"]

        if browser != 'Unknown' and osystem != 'Unknown':
            browser_chart[browser] = browser_chart.get(browser, 0) + d["total"]
            os_chart[osystem] = os_chart.get(osystem, 0) + d["total"]
            platform_chart[platform] = platform_chart.get(platform, 0) + d["total"]
            bot_chart[_("Human")] = bot_chart.get(_("Human"), 0) + d["total"]
            human_ua_table.append(
                {
                    "user_agent": d["user_agent"],
                    "total": d["total"],
                }
            )
        else:
            bot_chart[_("Bot")] = bot_chart.get(_("Bot"), 0) + d["total"]
            bot_ua_table.append(
                {
                    "user_agent": d["user_agent"],
                    "total": d["total"],
                }
            )

    bot_ua_table.sort(key=lambda ua: ua["total"], reverse=True)
    human_ua_table.sort(key=lambda ua: ua["total"], reverse=True)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Redaction")
    ctx['scope'] = scope
    ctx['platform_chart'] = dict(sorted(platform_chart.items(), key=lambda item: item[1], reverse=True))
    ctx['browser_chart'] = dict(sorted(browser_chart.items(), key=lambda item: item[1], reverse=True))
    ctx['os_chart'] = dict(sorted(os_chart.items(), key=lambda item: item[1], reverse=True))
    ctx['bot_chart'] = dict(sorted(bot_chart.items(), key=lambda item: item[1], reverse=True))
    ctx['bot_ua_table'] = bot_ua_table
    ctx['human_ua_table'] = human_ua_table
    ctx['header'] = header
    ctx['time_span_info'] = time_span_info
    ctx['aside_menu_items'] = get_side_menu('analytics', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('redaction', request.user)
    return render(request, "redaction_analytics.html", ctx)
