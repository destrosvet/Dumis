from . import utils, forms, models
from django import __version__ as django_version
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.models import User, Group, Permission
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.contrib.contenttypes.models import ContentType
from django.db.models import ProtectedError
from django.http import HttpResponse
from django.utils.dateparse import parse_datetime
from django.utils.translation import gettext_lazy as _
from django.utils.translation import gettext as gt
from django.utils.safestring import mark_safe
from django.urls import reverse
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
import sys
from .permissions import (
    svjis_view_admin_menu,
    svjis_edit_admin_users,
    svjis_edit_admin_groups,
    svjis_edit_admin_preferences,
    svjis_edit_admin_custom_fields,
    svjis_edit_admin_company,
    svjis_edit_admin_building,
    svjis_edit_admin_project_statuses,
    svjis_edit_admin_tags,
)
from svjis import __homepage_url__, __repository_url__, __issues_url__, __translations_url__


# Variables available in the mail template preferences. The order matters only
# for the data migration that converts the old positional {} placeholders.
MAIL_TEMPLATE_VARS = {
    'mail.template.lost.password': [
        ('message', _('Text with the username and the new password')),
    ],
    'mail.template.article.notification': [
        ('link', _('Link to the article')),
    ],
    'mail.template.comment.notification': [
        ('author', _('Comment author')),
        ('link', _('Link to the commented article')),
        ('comment', _('Comment body')),
    ],
    'mail.template.fault.notification': [
        ('author', _('Fault reporter')),
        ('link', _('Link to the fault')),
        ('description', _('Fault description')),
    ],
    'mail.template.fault.comment.notification': [
        ('author', _('Comment author')),
        ('link', _('Link to the fault')),
        ('comment', _('Comment body')),
    ],
    'mail.template.fault.assigned': [
        ('assignor', _('User who assigned the fault to you')),
        ('link', _('Link to the fault')),
        ('description', _('Fault description')),
    ],
    'mail.template.fault.closed': [
        ('user', _('User who closed the fault')),
        ('link', _('Link to the fault')),
        ('description', _('Fault description')),
    ],
    'mail.template.fault.reopened': [
        ('user', _('User who reopened the fault')),
        ('link', _('Link to the fault')),
        ('description', _('Fault description')),
    ],
    'mail.template.project.notification': [
        ('author', _('Project creator')),
        ('link', _('Link to the project')),
        ('description', _('Project description')),
    ],
    'mail.template.project.comment.notification': [
        ('author', _('Comment author')),
        ('link', _('Link to the project')),
        ('comment', _('Comment body')),
    ],
    'mail.template.project.assigned': [
        ('assignor', _('User who assigned the project to you')),
        ('link', _('Link to the project')),
        ('description', _('Project description')),
    ],
    'mail.template.project.status.changed': [
        ('user', _('User who changed the status')),
        ('link', _('Link to the project')),
        ('old_status', _('Previous status')),
        ('new_status', _('New status')),
    ],
}


def validate_template_placeholders(value, allowed):
    import re

    bad = []
    for m in re.finditer(r'\{[^{}]*\}', value):
        name = m.group(0)[1:-1]
        if name not in allowed:
            bad.append(m.group(0))
    return bad


# Custom fields
#####################


def get_custom_fields_context(model_cls, instance):
    ct = ContentType.objects.get_for_model(model_cls)
    defs = models.CustomFieldDefinition.objects.filter(content_type=ct).order_by('name')
    values = {}
    if instance and instance.pk:
        values = {
            v.field_definition_id: v.value
            for v in models.CustomFieldValue.objects.filter(content_type=ct, object_id=instance.pk)
        }
    return [{'definition': d, 'field_name': f'customfield_{d.id}', 'value': values.get(d.id, '')} for d in defs]


def save_custom_fields(request, instance):
    D = models.CustomFieldDefinition
    ct = ContentType.objects.get_for_model(instance)
    errors = []
    for d in D.objects.filter(content_type=ct):
        field_name = f'customfield_{d.id}'
        if d.field_type == D.TYPE_BOOLEAN:
            raw = '1' if request.POST.get(field_name) else ''
        else:
            raw = request.POST.get(field_name, '').strip()

        if raw and d.field_type == D.TYPE_NUMBER:
            try:
                float(raw.replace(',', '.'))
            except ValueError:
                errors.append(_('Field "{}" must be a number.').format(d.name))
                continue
        elif raw and d.field_type == D.TYPE_ENUM and raw not in d.enum_choices():
            errors.append(_('Field "{}" has an invalid value.').format(d.name))
            continue
        elif raw and d.field_type == D.TYPE_DATETIME and parse_datetime(raw) is None:
            errors.append(_('Field "{}" must be a valid date/time.').format(d.name))
            continue

        if raw:
            models.CustomFieldValue.objects.update_or_create(
                field_definition=d, content_type=ct, object_id=instance.pk, defaults={'value': raw}
            )
        else:
            models.CustomFieldValue.objects.filter(field_definition=d, content_type=ct, object_id=instance.pk).delete()
    return errors


# Tags
#####################


def get_tags_context(model_cls, instance):
    ct = ContentType.objects.get_for_model(model_cls)
    selected_ids = set()
    if instance and instance.pk:
        selected_ids = set(
            models.TaggedItem.objects.filter(content_type=ct, object_id=instance.pk).values_list('tag_id', flat=True)
        )
    return {
        'all_tags': models.Tag.objects.all(),
        'selected_tag_ids': selected_ids,
        'color_choices': models.BADGE_COLOR_CHOICES,
    }


def save_tags(request, instance):
    ct = ContentType.objects.get_for_model(instance)
    posted_ids = {int(pk) for pk in request.POST.getlist('tags') if pk.isdigit()}
    models.TaggedItem.objects.filter(content_type=ct, object_id=instance.pk).exclude(tag_id__in=posted_ids).delete()
    existing_ids = set(
        models.TaggedItem.objects.filter(content_type=ct, object_id=instance.pk).values_list('tag_id', flat=True)
    )
    for tag_id in posted_ids - existing_ids:
        models.TaggedItem.objects.get_or_create(tag_id=tag_id, content_type=ct, object_id=instance.pk)


def get_side_menu(active_item, user):
    side_menu = [
        {
            'perms': svjis_edit_admin_company,
            'description': _("Company"),
            'link': reverse(admin_company_edit_view),
            'active': True if active_item == 'company' else False,
        },
        {
            'perms': svjis_edit_admin_company,
            'description': _("Board") + f' ({models.Board.objects.count()})',
            'link': reverse(admin_board_view),
            'active': True if active_item == 'board' else False,
        },
        {
            'perms': svjis_edit_admin_building,
            'description': _("Building"),
            'link': reverse(admin_building_edit_view),
            'active': True if active_item == 'building' else False,
        },
        {
            'perms': svjis_edit_admin_building,
            'description': _("Entrances") + f' ({models.BuildingEntrance.objects.count()})',
            'link': reverse(admin_entrance_view),
            'active': True if active_item == 'entrances' else False,
        },
        {
            'perms': svjis_edit_admin_building,
            'description': _("Building units") + f' ({models.BuildingUnit.objects.count()})',
            'link': reverse(admin_building_unit_view),
            'active': True if active_item == 'units' else False,
        },
        {
            'perms': svjis_edit_admin_users,
            'description': _("Users") + f' ({User.objects.filter(is_active=True).count()})',
            'link': reverse(admin_user_view),
            'active': True if active_item == 'users' else False,
        },
        {
            'perms': svjis_edit_admin_groups,
            'description': _("Groups"),
            'link': reverse(admin_group_view),
            'active': True if active_item == 'groups' else False,
        },
        {
            'perms': svjis_edit_admin_preferences,
            'description': _("E-Mail templates"),
            'link': reverse(admin_preferences_view),
            'active': True if active_item == 'preferences' else False,
        },
        {
            'perms': svjis_edit_admin_custom_fields,
            'description': _("Custom fields"),
            'link': reverse(admin_custom_field_view),
            'active': True if active_item == 'custom_fields' else False,
        },
        {
            'perms': svjis_edit_admin_project_statuses,
            'description': _("Project statuses"),
            'link': reverse(admin_project_status_view),
            'active': True if active_item == 'project_statuses' else False,
        },
        {
            'perms': svjis_edit_admin_tags,
            'description': _("Tags"),
            'link': reverse(admin_tag_view),
            'active': True if active_item == 'tags' else False,
        },
        {
            'perms': svjis_view_admin_menu,
            'description': _("Waiting messages") + f' ({models.MessageQueue.objects.filter(status=0).count()})',
            'link': reverse(admin_messages_view),
            'active': True if active_item == 'messages' else False,
        },
        {
            'perms': svjis_view_admin_menu,
            'description': _("About application"),
            'link': reverse(admin_about_view),
            'active': True if active_item == 'about' else False,
        },
    ]
    return [x for x in side_menu if x['perms'] is None or user.has_perm(x['perms'])]


# Administration - Company
@permission_required(svjis_edit_admin_company)
@require_GET
def admin_company_edit_view(request):
    instance, _created = models.Company.objects.get_or_create(pk=1)
    form = forms.CompanyForm(instance=instance)
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['aside_menu_items'] = get_side_menu('company', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_company_edit.html", ctx)


@permission_required(svjis_edit_admin_company)
@require_POST
def admin_company_save_view(request):
    instance, _created = models.Company.objects.get_or_create(pk=1)
    form = forms.CompanyForm(request.POST, request.FILES, instance=instance)
    if form.is_valid:
        form.save()
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
    return redirect(admin_company_edit_view)


@permission_required(svjis_edit_admin_company)
@require_GET
def admin_company_remove_logo_view(request):
    instance, _created = models.Company.objects.get_or_create(pk=1)
    instance.header_picture = None
    instance.save()
    messages.info(request, _('Saved'))
    return redirect(admin_company_edit_view)


# Administration - Board


@permission_required(svjis_edit_admin_company)
@require_GET
def admin_board_view(request):
    board_list = models.Board.objects.select_related('member', 'member__userprofile')
    rows = []
    for b in board_list:
        member_name = f"{b.member.userprofile.salutation} {b.member.first_name} {b.member.last_name}".strip()
        rows.append(
            {
                'url': reverse('admin_board_edit', args=[b.pk]),
                'cells': {'member': member_name, 'position': b.position},
                'actions': [
                    {'url': reverse('admin_board_edit', args=[b.pk]), 'icon': 'edit', 'label': _('Edit')},
                    {
                        'url': reverse('admin_board_delete', args=[b.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete')} {member_name} ?",
                    },
                ],
            }
        )
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('board', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = board_list
    ctx['columns'] = [
        {'key': 'member', 'label': _('Member')},
        {'key': 'position', 'label': _('Position')},
    ]
    ctx['rows'] = rows
    ctx['table_id'] = 'board-table'
    ctx['desc_id'] = 'tbl-desc'
    ctx['list_search'] = True
    return render(request, "admin_board.html", ctx)


@permission_required(svjis_edit_admin_company)
@require_GET
def admin_board_edit_view(request, pk):
    if pk:
        i = get_object_or_404(models.Board, pk=pk)
        form = forms.BoardForm(instance=i)
    else:
        i = None
        form = forms.BoardForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['custom_fields'] = get_custom_fields_context(models.Board, i)
    ctx['aside_menu_items'] = get_side_menu('board', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_board_edit.html", ctx)


@permission_required(svjis_edit_admin_company)
@require_POST
def admin_board_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.BoardForm(request.POST)
    else:
        instance = get_object_or_404(models.Board, pk=pk)
        form = forms.BoardForm(request.POST, instance=instance)
    if form.is_valid:
        obj = form.save(commit=False)
        obj.company = models.Company.objects.get(pk=1)
        obj.save()
        for error in save_custom_fields(request, obj):
            messages.error(request, error)
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
    return redirect(admin_board_view)


@permission_required(svjis_edit_admin_company)
@require_GET
def admin_board_delete_view(request, pk):
    obj = get_object_or_404(models.Board, pk=pk)
    obj.delete()
    return redirect(admin_board_view)


# Administration - Building
@permission_required(svjis_edit_admin_building)
@require_GET
def admin_building_edit_view(request):
    instance, _created = models.Building.objects.get_or_create(pk=1)
    form = forms.BuildingForm(instance=instance)
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['aside_menu_items'] = get_side_menu('building', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_building_edit.html", ctx)


@permission_required(svjis_edit_admin_building)
@require_POST
def admin_building_save_view(request):
    instance, _created = models.Building.objects.get_or_create(pk=1)
    form = forms.BuildingForm(request.POST, instance=instance)
    if form.is_valid:
        form.save()
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
    return redirect(admin_building_edit_view)


# Administration - BuildingEntrance
@permission_required(svjis_edit_admin_building)
@require_GET
def admin_entrance_view(request):
    entrance_list = models.BuildingEntrance.objects.all()
    rows = []
    for e in entrance_list:
        rows.append(
            {
                'url': reverse('admin_entrance_edit', args=[e.pk]),
                'cells': {'description': e.description, 'address': e.address},
                'actions': [
                    {'url': reverse('admin_entrance_edit', args=[e.pk]), 'icon': 'edit', 'label': _('Edit')},
                    {
                        'url': reverse('admin_entrance_delete', args=[e.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete')} {e.description} ?",
                    },
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('entrances', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = entrance_list
    ctx['columns'] = [
        {'key': 'description', 'label': _('Description')},
        {'key': 'address', 'label': _('Address')},
    ]
    ctx['rows'] = rows
    ctx['table_id'] = 'entrance-table'
    ctx['desc_id'] = 'tbl-desc'
    ctx['list_search'] = True
    return render(request, "admin_entrance.html", ctx)


@permission_required(svjis_edit_admin_building)
@require_GET
def admin_entrance_edit_view(request, pk):
    if pk:
        i = get_object_or_404(models.BuildingEntrance, pk=pk)
        form = forms.BuildingEntranceForm(instance=i)
    else:
        i = None
        form = forms.BuildingEntranceForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['custom_fields'] = get_custom_fields_context(models.BuildingEntrance, i)
    ctx['aside_menu_items'] = get_side_menu('entrances', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_entrance_edit.html", ctx)


@permission_required(svjis_edit_admin_building)
@require_POST
def admin_entrance_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.BuildingEntranceForm(request.POST)
    else:
        instance = get_object_or_404(models.BuildingEntrance, pk=pk)
        form = forms.BuildingEntranceForm(request.POST, instance=instance)
    if form.is_valid:
        obj = form.save(commit=False)
        obj.building, _created = models.Building.objects.get_or_create(pk=1)
        obj.save()
        for error in save_custom_fields(request, obj):
            messages.error(request, error)
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")

    return redirect(admin_entrance_view)


@permission_required(svjis_edit_admin_building)
@require_GET
def admin_entrance_delete_view(request, pk):
    obj = get_object_or_404(models.BuildingEntrance, pk=pk)
    obj.delete()
    return redirect(admin_entrance_view)


# Administration - BuildingUnit
@permission_required(svjis_edit_admin_building)
@require_GET
def admin_building_unit_view(request):
    unit_list = models.BuildingUnit.objects.select_related('type', 'entrance')
    type_list = models.BuildingUnitType.objects.all()
    entrance_list = models.BuildingEntrance.objects.all()

    type_filter = int(request.GET.get('type_filter', 0))
    if type_filter != 0:
        unit_list = unit_list.filter(type_id=type_filter)

    entrance_filter = int(request.GET.get('entrance_filter', 0))
    if entrance_filter != 0:
        unit_list = unit_list.filter(entrance_id=entrance_filter)

    rows = []
    for u in unit_list.order_by('id'):
        unit_desc = f"{u.type.description} {u.description}"
        actions = [
            {'url': reverse('admin_building_unit_edit', args=[u.pk]), 'icon': 'edit', 'label': _('Edit')},
            {'url': reverse('admin_building_unit_owners', args=[u.pk]), 'icon': 'user', 'label': _('Owners')},
        ]
        if not u.owners.exists():
            actions.append(
                {
                    'url': reverse('admin_user_edit', args=[0]) + f'?unit={u.pk}&role=owner',
                    'icon': 'user',
                    'label': _('Add owner'),
                }
            )
        actions.append(
            {
                'url': reverse('admin_user_edit', args=[0]) + f'?unit={u.pk}&role=tenant',
                'icon': 'user',
                'label': _('Add tenant'),
            }
        )
        actions.append(
            {
                'url': reverse('admin_building_unit_delete', args=[u.pk]),
                'icon': 'delete',
                'label': _('Delete'),
                'danger': True,
                'confirm': f"{_('Do you want to delete')} {unit_desc} ?",
            }
        )
        rows.append(
            {
                'url': reverse('admin_building_unit_edit', args=[u.pk]),
                'cells': {
                    'type': u.type.description,
                    'entrance': u.entrance.description if u.entrance else '',
                    'registration_id': u.registration_id,
                    'description': u.description,
                    'numerator': u.numerator,
                    'denominator': u.denominator,
                },
                'actions': actions,
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('units', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = unit_list.order_by('id')
    ctx['type_list'] = type_list
    ctx['type_filter'] = type_filter
    ctx['entrance_list'] = entrance_list
    ctx['entrance_filter'] = entrance_filter
    ctx['columns'] = [
        {'key': 'type', 'label': _('Type')},
        {'key': 'entrance', 'label': _('Entrance')},
        {'key': 'registration_id', 'label': _('Registration Id')},
        {'key': 'description', 'label': _('Description'), 'hide_on_mobile': True},
        {'key': 'numerator', 'label': _('Numerator'), 'hide_on_mobile': True},
        {'key': 'denominator', 'label': _('Denominator'), 'hide_on_mobile': True},
    ]
    ctx['rows'] = rows
    ctx['table_id'] = 'building-unit-table'
    ctx['desc_id'] = 'tbl-desc'
    ctx['list_search'] = True
    return render(request, "admin_building_unit.html", ctx)


@permission_required(svjis_edit_admin_building)
@require_GET
def admin_building_unit_edit_view(request, pk):
    if pk:
        i = get_object_or_404(models.BuildingUnit, pk=pk)
        form = forms.BuildingUnitForm(instance=i)
    else:
        i = None
        form = forms.BuildingUnitForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['custom_fields'] = get_custom_fields_context(models.BuildingUnit, i)
    ctx['aside_menu_items'] = get_side_menu('units', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_building_unit_edit.html", ctx)


@permission_required(svjis_edit_admin_building)
@require_POST
def admin_building_unit_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.BuildingUnitForm(request.POST)
    else:
        instance = get_object_or_404(models.BuildingUnit, pk=pk)
        form = forms.BuildingUnitForm(request.POST, instance=instance)
    if form.is_valid:
        obj = form.save(commit=False)
        obj.building, _created = models.Building.objects.get_or_create(pk=1)
        obj.save()
        for error in save_custom_fields(request, obj):
            messages.error(request, error)
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")

    qs = f"#unit_{obj.pk}"
    return redirect(reverse(admin_building_unit_view) + qs)


@permission_required(svjis_edit_admin_building)
@require_GET
def admin_building_unit_delete_view(request, pk):
    obj = get_object_or_404(models.BuildingUnit, pk=pk)
    obj.delete()
    return redirect(admin_building_unit_view)


@permission_required(svjis_edit_admin_building)
@require_GET
def admin_building_unit_owners_view(request, pk):
    bu = get_object_or_404(models.BuildingUnit, pk=pk)
    user_list = User.objects.filter(is_active=True).order_by('last_name', 'first_name')
    owner_candidates = user_list.exclude(id__in=bu.owners.values('id'))
    tenant_candidates = user_list.exclude(id__in=bu.tenants.values('id'))

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['bu'] = bu
    ctx['pk'] = pk
    ctx['owner_list'] = bu.owners
    ctx['tenant_list'] = bu.tenants
    ctx['user_list'] = user_list
    ctx['owner_candidates'] = owner_candidates
    ctx['tenant_candidates'] = tenant_candidates
    ctx['aside_menu_items'] = get_side_menu('units', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['breadcrumbs'] = [
        {'label': _('Administration'), 'url': reverse(admin_company_edit_view)},
        {'label': _('Building units'), 'url': reverse(admin_building_unit_view)},
        {'label': f'{bu.type.description} - {bu.registration_id} - {bu.description}'},
    ]
    return render(request, "admin_building_unit_owners_edit.html", ctx)


@permission_required(svjis_edit_admin_building)
@require_POST
def admin_building_unit_owners_save_view(request):
    pk = int(request.POST['pk'])
    user_id = int(request.POST['owner_id'])
    role = request.POST.get('role', models.BuildingUnitUser.ROLE_OWNER)

    if user_id > 0 and role in (models.BuildingUnitUser.ROLE_OWNER, models.BuildingUnitUser.ROLE_TENANT):
        bu = get_object_or_404(models.BuildingUnit, pk=pk)
        u = get_object_or_404(User, pk=user_id)
        models.BuildingUnitUser.objects.get_or_create(building_unit=bu, user=u, role=role)

    return redirect(admin_building_unit_owners_view, pk=pk)


@permission_required(svjis_edit_admin_building)
@require_GET
def admin_building_unit_owners_delete_view(request, pk, role, user):
    bu = get_object_or_404(models.BuildingUnit, pk=pk)
    u = get_object_or_404(User, pk=user)
    models.BuildingUnitUser.objects.filter(building_unit=bu, user=u, role=role).delete()
    return redirect(admin_building_unit_owners_view, pk=pk)


@permission_required(svjis_edit_admin_building)
@require_GET
def admin_building_unit_export_to_excel_view(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Building_Units.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = gt("Building units")

    # Add headers
    headers = [
        gt("Type"),
        gt("Entrance"),
        gt("Registration Id"),
        gt("Description"),
        gt("Numerator"),
        gt("Denominator"),
    ]
    ws.append(headers)

    header_st = utils.get_worksheet_header_style()

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in rows:
            cell.style = header_st

    # Add data from the model
    unit_list = models.BuildingUnit.objects.select_related('type', 'entrance').order_by('id')
    for u in unit_list:
        unit_entrance = u.entrance.description if u.entrance else ''
        ws.append([u.type.description, unit_entrance, u.registration_id, u.description, u.numerator, u.denominator])

    utils.adjust_worksheet_columns_width(ws)

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response


# Administration - User
@permission_required(svjis_edit_admin_users)
@require_GET
def admin_user_view(request):
    deactivated_users = request.GET.get('deactivated_users', False) == 'on'
    user_list = User.objects.filter(is_active=not deactivated_users)
    group_filter = int(request.GET.get('group_filter', 0))
    if group_filter != 0:
        g = Group.objects.filter(pk=group_filter)
        user_list = user_list.filter(groups__in=g)
    group_list = Group.objects.all()

    rows = []
    for u in user_list.order_by('last_name', 'first_name'):
        last_login = u.last_login.strftime('%d.%m.%Y %H:%M') if u.last_login else ''
        rows.append(
            {
                'url': reverse('admin_user_detail', args=[u.pk]),
                'cells': {
                    'last_name': u.last_name,
                    'first_name': u.first_name,
                    'email': mark_safe(
                        f'<a href="mailto:{u.email}" onclick="event.stopPropagation()">{u.email}</a>'
                    ),
                    'last_login': last_login,
                },
                'actions': [
                    {'url': reverse('admin_user_edit', args=[u.pk]), 'icon': 'edit', 'label': _('Edit')},
                    {'url': reverse('admin_user_detail', args=[u.pk]), 'icon': 'view', 'label': _('Detail')},
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('users', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = user_list.order_by('last_name', 'first_name')
    ctx['group_list'] = group_list.order_by('name')
    ctx['group_filter'] = group_filter
    ctx['deactivated_users'] = deactivated_users
    ctx['columns'] = [
        {'key': 'last_name', 'label': _('Last name')},
        {'key': 'first_name', 'label': _('First name')},
        {'key': 'email', 'label': _('E-Mail'), 'hide_on_mobile': True},
        {'key': 'last_login', 'label': _('Last login'), 'hide_on_mobile': True},
    ]
    ctx['rows'] = rows
    ctx['table_id'] = 'user-table'
    ctx['desc_id'] = 'users'
    ctx['list_search'] = True
    return render(request, "admin_user.html", ctx)


@permission_required(svjis_edit_admin_users)
@require_GET
def admin_user_detail_view(request, pk):
    user = get_object_or_404(User, pk=pk)
    profile, _created = models.UserProfile.objects.get_or_create(user=user)
    user_group_list = Group.objects.filter(user__id=user.id)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['usr'] = user
    ctx['profile'] = profile
    ctx['user_group_list'] = user_group_list
    ctx['custom_fields'] = get_custom_fields_context(User, user)
    ctx['aside_menu_items'] = get_side_menu('users', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['breadcrumbs'] = [
        {'label': _('Administration'), 'url': reverse(admin_company_edit_view)},
        {'label': _('Users'), 'url': reverse(admin_user_view)},
        {'label': f'{user.first_name} {user.last_name}'},
    ]
    return render(request, "admin_user_detail.html", ctx)


@permission_required(svjis_edit_admin_users)
@require_GET
def admin_user_edit_view(request, pk):
    if pk:
        instance = get_object_or_404(User, pk=pk)
        uform = forms.UserForm(instance=instance)
        pinstance, _created = models.UserProfile.objects.get_or_create(user=instance)
        pform = forms.UserProfileForm(instance=pinstance)
    else:
        instance = User
        uform = forms.UserForm()
        pform = forms.UserProfileForm()

    unit_pk = int(request.GET.get('unit', 0))
    unit_role = request.GET.get('role', '')
    unit = None
    if unit_pk and unit_role in (models.BuildingUnitUser.ROLE_OWNER, models.BuildingUnitUser.ROLE_TENANT):
        unit = get_object_or_404(models.BuildingUnit, pk=unit_pk)

    group_list = []
    user_group_list = []
    if pk:
        user_group_list = Group.objects.filter(user__id=instance.id)
    for g in Group.objects.all().order_by('name'):
        item = {'name': g.name, 'checked': g in user_group_list}
        group_list.append(item)

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['uform'] = uform
    ctx['pform'] = pform
    ctx['group_list'] = group_list
    ctx['pk'] = pk
    ctx['unit'] = unit
    ctx['unit_role'] = unit_role
    ctx['custom_fields'] = get_custom_fields_context(User, instance if pk else None)
    ctx['aside_menu_items'] = get_side_menu('users', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_user_edit.html", ctx)


@permission_required(svjis_edit_admin_users)
@require_POST
def admin_user_save_view(request):
    pk = int(request.POST['pk'])
    unit_pk = int(request.POST.get('unit', 0))
    unit_role = request.POST.get('unit_role', '')
    if pk:
        user_i = get_object_or_404(User, pk=pk)
        user_form = forms.UserForm(request.POST, instance=user_i)
        user_profile_form = forms.UserProfileForm(request.POST, instance=user_i.userprofile)
    else:
        user_form = forms.UserForm(request.POST)
        user_profile_form = forms.UserProfileForm(request.POST)

    if user_form.is_valid() and user_profile_form.is_valid():
        u = user_form.save()
        u.userprofile = user_profile_form.instance
        u.userprofile.save()

        password = request.POST.get('password', '')
        if password != '':
            u.password = make_password(password)
            u.save()

        send_credentials = request.POST.get('send_credentials', False) == 'on'
        if send_credentials:
            utils.send_new_password(u)
            messages.info(request, _('Credentials has been sent by e-mail'))

        # Set groups
        user_group_list = Group.objects.filter(user__id=u.id)
        for g in Group.objects.all():
            group_set = request.POST.get(g.name, False)
            if group_set and g not in user_group_list:
                u.groups.add(g)
            if not group_set and g in user_group_list:
                u.groups.remove(g)

        # Assign the user to a building unit (owner/tenant) when requested
        if unit_pk and unit_role in (models.BuildingUnitUser.ROLE_OWNER, models.BuildingUnitUser.ROLE_TENANT):
            bu = get_object_or_404(models.BuildingUnit, pk=unit_pk)
            models.BuildingUnitUser.objects.get_or_create(building_unit=bu, user=u, role=unit_role)

        for error in save_custom_fields(request, u):
            messages.error(request, error)

    else:
        for error in user_form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
        for error in user_profile_form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
        return redirect(reverse('admin_user_edit', kwargs={'pk': pk}))

    messages.info(request, _('Saved'))
    if unit_pk:
        return redirect(admin_building_unit_owners_view, pk=unit_pk)
    return redirect(admin_user_detail_view, pk=u.pk)


@permission_required(svjis_edit_admin_users)
@require_GET
def admin_user_owns_view(request, pk):
    u = get_object_or_404(User, pk=pk)
    membership_list = models.BuildingUnitUser.objects.filter(user=u).select_related(
        'building_unit', 'building_unit__type'
    )
    owner_list = membership_list.filter(role=models.BuildingUnitUser.ROLE_OWNER)
    tenant_list = membership_list.filter(role=models.BuildingUnitUser.ROLE_TENANT)
    bu_list = models.BuildingUnit.objects.exclude(id__in=u.building_units.values_list("id", flat=True)).order_by("id")
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['u'] = u
    ctx['pk'] = pk
    ctx['membership_list'] = membership_list
    ctx['owner_list'] = owner_list
    ctx['tenant_list'] = tenant_list
    ctx['bu_list'] = bu_list
    ctx['aside_menu_items'] = get_side_menu('users', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['breadcrumbs'] = [
        {'label': _('Administration'), 'url': reverse(admin_company_edit_view)},
        {'label': _('Users'), 'url': reverse(admin_user_view)},
        {'label': f'{u.first_name} {u.last_name}', 'url': reverse(admin_user_detail_view, args=[u.pk])},
        {'label': _('Modify units')},
    ]
    return render(request, "admin_user_ownes_edit.html", ctx)


@permission_required(svjis_edit_admin_users)
@require_POST
def admin_user_owns_save_view(request):
    pk = int(request.POST['pk'])
    unit_id = int(request.POST['unit_id'])
    role = request.POST.get('role', models.BuildingUnitUser.ROLE_OWNER)

    if unit_id > 0 and role in (models.BuildingUnitUser.ROLE_OWNER, models.BuildingUnitUser.ROLE_TENANT):
        u = get_object_or_404(User, pk=pk)
        bu = get_object_or_404(models.BuildingUnit, pk=unit_id)
        models.BuildingUnitUser.objects.get_or_create(building_unit=bu, user=u, role=role)

    return redirect(admin_user_owns_view, pk=pk)


@permission_required(svjis_edit_admin_users)
@require_GET
def admin_user_owns_delete_view(request, pk, unit, role):
    u = get_object_or_404(User, pk=pk)
    bu = get_object_or_404(models.BuildingUnit, pk=unit)
    models.BuildingUnitUser.objects.filter(building_unit=bu, user=u, role=role).delete()
    return redirect(admin_user_owns_view, pk=pk)


@permission_required(svjis_edit_admin_users)
@require_GET
def admin_user_export_to_excel_view(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Users.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = gt("Users")

    # Add headers
    headers = [
        gt("Id"),
        gt("Salutation"),
        gt("First name"),
        gt("Last name"),
        gt("Address"),
        gt("City"),
        gt("Post code"),
        gt("Country"),
        gt("Phone"),
        gt("Email address"),
        gt("Username"),
    ]
    ws.append(headers)

    header_st = utils.get_worksheet_header_style()

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in rows:
            cell.style = header_st

    # Add data from the model
    user_list = User.objects.filter(is_active=True).select_related('userprofile').order_by('last_name', 'first_name')
    for u in user_list:
        if hasattr(u, 'userprofile'):
            ws.append(
                [
                    u.id,
                    u.userprofile.salutation,
                    u.first_name,
                    u.last_name,
                    u.userprofile.address,
                    u.userprofile.city,
                    u.userprofile.post_code,
                    u.userprofile.country,
                    u.userprofile.phone,
                    u.email,
                    u.username,
                ]
            )

    utils.adjust_worksheet_columns_width(ws)

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response


# Administration - Group
@permission_required(svjis_edit_admin_groups)
@require_GET
def admin_group_view(request):
    group_list = Group.objects.all()
    rows = []
    for g in group_list.order_by('name'):
        rows.append(
            {
                'url': reverse('admin_group_edit', args=[g.pk]),
                'cells': {'name': g.name},
                'actions': [
                    {'url': reverse('admin_group_edit', args=[g.pk]), 'icon': 'edit', 'label': _('Edit')},
                    {
                        'url': reverse('admin_user') + f'?group_filter={g.pk}',
                        'icon': 'user',
                        'label': _('List of users'),
                    },
                    {
                        'url': reverse('admin_group_delete', args=[g.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete')} {g.name} ?",
                    },
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('groups', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = group_list.order_by('name')
    ctx['columns'] = [{'key': 'name', 'label': _('Name')}]
    ctx['rows'] = rows
    ctx['table_id'] = 'group-table'
    ctx['desc_id'] = 'groups'
    ctx['list_search'] = True
    return render(request, "admin_group.html", ctx)


@permission_required(svjis_edit_admin_groups)
@require_GET
def admin_group_edit_view(request, pk):
    if pk:
        i = get_object_or_404(Group, pk=pk)
        form = forms.GroupEditForm(instance=i)
    else:
        i = Group
        form = forms.GroupEditForm

    permission_list = [
        {
            'name': 'Articles',
            'perms': [{'codename': 'svjis_add_article_comment'}, {'codename': 'svjis_answer_survey'}],
        },
        {'name': 'Contact', 'perms': [{'codename': 'svjis_view_phonelist'}]},
        {'name': 'Personal settings', 'perms': [{'codename': 'svjis_view_personal_menu'}]},
        {
            'name': 'Redaction',
            'perms': [
                {'codename': 'svjis_view_redaction_menu'},
                {'codename': 'svjis_edit_article'},
                {'codename': 'svjis_edit_article_news'},
                {'codename': 'svjis_edit_useful_link'},
                {'codename': 'svjis_edit_survey'},
                {'codename': 'svjis_edit_article_menu'},
            ],
        },
        {
            'name': 'Fault reporting',
            'perms': [
                {'codename': 'svjis_view_fault_menu'},
                {'codename': 'svjis_add_fault_comment'},
                {'codename': 'svjis_fault_reporter'},
                {'codename': 'svjis_fault_resolver'},
            ],
        },
        {'name': 'Adverts', 'perms': [{'codename': 'svjis_view_adverts_menu'}, {'codename': 'svjis_add_advert'}]},
        {
            'name': 'Projects',
            'perms': [
                {'codename': 'svjis_view_project_menu'},
                {'codename': 'svjis_add_project'},
                {'codename': 'svjis_manage_projects'},
                {'codename': 'svjis_add_project_comment'},
            ],
        },
        {
            'name': 'Administration',
            'perms': [
                {'codename': 'svjis_view_admin_menu'},
                {'codename': 'svjis_edit_admin_company'},
                {'codename': 'svjis_edit_admin_building'},
                {'codename': 'svjis_edit_admin_users'},
                {'codename': 'svjis_edit_admin_groups'},
                {'codename': 'svjis_edit_admin_preferences'},
                {'codename': 'svjis_edit_admin_project_statuses'},
                {'codename': 'svjis_edit_admin_tags'},
            ],
        },
    ]
    group_perm_list = []
    if pk:
        group_perm_list = Permission.objects.filter(group__id=i.id)
    for m in permission_list:
        for p in m['perms']:
            perm = Permission.objects.filter(codename=p['codename']).first()
            p['name'] = perm.name
            p['checked'] = perm in group_perm_list

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['instance'] = i
    ctx['permission_list'] = permission_list
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('groups', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_group_edit.html", ctx)


@permission_required(svjis_edit_admin_groups)
@require_POST
def admin_group_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.GroupEditForm(request.POST)
    else:
        instance = get_object_or_404(Group, pk=pk)
        form = forms.GroupEditForm(request.POST, instance=instance)
    if form.is_valid:
        instance = form.save()

        # Set permissions
        group_perm_list = Permission.objects.filter(group__id=instance.id)
        for p in Permission.objects.all():
            if p.codename.startswith('svjis_'):
                perm_set = request.POST.get(p.codename, False) == 'on'
                if perm_set and p not in group_perm_list:
                    instance.permissions.add(p)
                if not perm_set and p in group_perm_list:
                    instance.permissions.remove(p)
        messages.info(request, _('Saved'))
    else:
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")

    return redirect(admin_group_view)


@permission_required(svjis_edit_admin_groups)
@require_GET
def admin_group_delete_view(request, pk):
    obj = get_object_or_404(Group, pk=pk)
    obj.delete()
    return redirect(admin_group_view)


# Administration - Preferences
@permission_required(svjis_edit_admin_preferences)
@require_GET
def admin_preferences_view(request):
    property_list = models.Preferences.objects.all()
    rows = []
    for p in property_list.order_by('key'):
        rows.append(
            {
                'url': reverse('admin_preferences_edit', args=[p.pk]),
                'cells': {'key': p.key, 'value': p.value},
                'actions': [
                    {'url': reverse('admin_preferences_edit', args=[p.pk]), 'icon': 'edit', 'label': _('Edit')},
                    {
                        'url': reverse('admin_preferences_delete', args=[p.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete')} {p.key} ?",
                    },
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('preferences', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = property_list.order_by('key')
    ctx['columns'] = [
        {'key': 'key', 'label': _('Key')},
        {'key': 'value', 'label': _('Value'), 'hide_on_mobile': True},
    ]
    ctx['rows'] = rows
    ctx['table_id'] = 'preferences-table'
    ctx['desc_id'] = 'preferences'
    ctx['list_search'] = True
    return render(request, "admin_preferences.html", ctx)


@permission_required(svjis_edit_admin_preferences)
@require_GET
def admin_preferences_edit_view(request, pk):
    if pk:
        i = get_object_or_404(models.Preferences, pk=pk)
        form = forms.PreferencesForm(instance=i)
    else:
        i = None
        form = forms.PreferencesForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['template_vars'] = MAIL_TEMPLATE_VARS.get(i.key) if i else None
    ctx['aside_menu_items'] = get_side_menu('preferences', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_preferences_edit.html", ctx)


@permission_required(svjis_edit_admin_preferences)
@require_POST
def admin_preferences_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.PreferencesForm(request.POST)
    else:
        instance = get_object_or_404(models.Preferences, pk=pk)
        form = forms.PreferencesForm(request.POST, instance=instance)

    if not form.is_valid():
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
        return redirect(admin_preferences_view)

    template_vars = MAIL_TEMPLATE_VARS.get(form.cleaned_data['key'])
    if template_vars is not None:
        allowed = [name for name, _ in template_vars]
        bad = validate_template_placeholders(form.cleaned_data['value'], allowed)
        if bad:
            allowed_text = ', '.join('{' + name + '}' for name in allowed)
            for item in bad:
                messages.error(
                    request,
                    _('Unknown placeholder "{}" in a mail template. Allowed placeholders: {}.').format(
                        item, allowed_text
                    ),
                )
            ctx = utils.get_context()
            ctx['aside_menu_name'] = _("Administration")
            ctx['form'] = form
            ctx['pk'] = pk
            ctx['template_vars'] = template_vars
            ctx['aside_menu_items'] = get_side_menu('preferences', request.user)
            ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
            return render(request, "admin_preferences_edit.html", ctx)

    form.save()
    messages.info(request, _('Saved'))
    return redirect(admin_preferences_view)


@permission_required(svjis_edit_admin_preferences)
@require_GET
def admin_preferences_delete_view(request, pk):
    obj = get_object_or_404(models.Preferences, pk=pk)
    obj.delete()
    return redirect(admin_preferences_view)


# Administration - Custom fields
@permission_required(svjis_edit_admin_custom_fields)
@require_GET
def admin_custom_field_view(request):
    definition_list = models.CustomFieldDefinition.objects.select_related('content_type')
    entity_labels = models.get_custom_field_model_labels()
    rows = []
    for d in definition_list:
        rows.append(
            {
                'url': reverse('admin_custom_field_edit', args=[d.pk]),
                'cells': {
                    'entity': entity_labels.get(d.content_type_id, str(d.content_type)),
                    'name': d.name,
                    'field_type': d.get_field_type_display(),
                },
                'actions': [
                    {'url': reverse('admin_custom_field_edit', args=[d.pk]), 'icon': 'edit', 'label': _('Edit')},
                    {
                        'url': reverse('admin_custom_field_delete', args=[d.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete')} {d.name} ?",
                    },
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('custom_fields', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = definition_list
    ctx['columns'] = [
        {'key': 'entity', 'label': _('Entity type')},
        {'key': 'name', 'label': _('Name')},
        {'key': 'field_type', 'label': _('Type'), 'hide_on_mobile': True},
    ]
    ctx['rows'] = rows
    ctx['table_id'] = 'custom-field-table'
    ctx['desc_id'] = 'custom-fields'
    ctx['list_search'] = True
    return render(request, "admin_custom_field.html", ctx)


@permission_required(svjis_edit_admin_custom_fields)
@require_GET
def admin_custom_field_edit_view(request, pk):
    if pk:
        i = get_object_or_404(models.CustomFieldDefinition, pk=pk)
        form = forms.CustomFieldDefinitionForm(instance=i)
    else:
        form = forms.CustomFieldDefinitionForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('custom_fields', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_custom_field_edit.html", ctx)


@permission_required(svjis_edit_admin_custom_fields)
@require_POST
def admin_custom_field_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.CustomFieldDefinitionForm(request.POST)
    else:
        instance = get_object_or_404(models.CustomFieldDefinition, pk=pk)
        form = forms.CustomFieldDefinitionForm(request.POST, instance=instance)

    if not form.is_valid():
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
        return redirect(admin_custom_field_view)

    form.save()
    messages.info(request, _('Saved'))
    return redirect(admin_custom_field_view)


@permission_required(svjis_edit_admin_custom_fields)
@require_GET
def admin_custom_field_delete_view(request, pk):
    obj = get_object_or_404(models.CustomFieldDefinition, pk=pk)
    obj.delete()
    return redirect(admin_custom_field_view)


# Administration - Project statuses
@permission_required(svjis_edit_admin_project_statuses)
@require_GET
def admin_project_status_view(request):
    status_list = models.ProjectStatus.objects.all()
    rows = []
    for s in status_list:
        rows.append(
            {
                'url': reverse('admin_project_status_edit', args=[s.pk]),
                'cells': {
                    'name': s.name,
                    'order': s.order,
                    'is_closed': _('Yes') if s.is_closed else _('No'),
                },
                'actions': [
                    {'url': reverse('admin_project_status_edit', args=[s.pk]), 'icon': 'edit', 'label': _('Edit')},
                    {
                        'url': reverse('admin_project_status_delete', args=[s.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete')} {s.name} ?",
                    },
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('project_statuses', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = status_list
    ctx['columns'] = [
        {'key': 'name', 'label': _('Name')},
        {'key': 'order', 'label': _('Order'), 'hide_on_mobile': True},
        {'key': 'is_closed', 'label': _('Closed state'), 'hide_on_mobile': True},
    ]
    ctx['rows'] = rows
    ctx['table_id'] = 'project-status-table'
    ctx['desc_id'] = 'project-statuses'
    ctx['list_search'] = True
    return render(request, "admin_project_status.html", ctx)


@permission_required(svjis_edit_admin_project_statuses)
@require_GET
def admin_project_status_edit_view(request, pk):
    if pk:
        i = get_object_or_404(models.ProjectStatus, pk=pk)
        form = forms.ProjectStatusForm(instance=i)
    else:
        form = forms.ProjectStatusForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('project_statuses', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_project_status_edit.html", ctx)


@permission_required(svjis_edit_admin_project_statuses)
@require_POST
def admin_project_status_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.ProjectStatusForm(request.POST)
    else:
        instance = get_object_or_404(models.ProjectStatus, pk=pk)
        form = forms.ProjectStatusForm(request.POST, instance=instance)

    if not form.is_valid():
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
        return redirect(admin_project_status_view)

    form.save()
    messages.info(request, _('Saved'))
    return redirect(admin_project_status_view)


@permission_required(svjis_edit_admin_project_statuses)
@require_GET
def admin_project_status_delete_view(request, pk):
    obj = get_object_or_404(models.ProjectStatus, pk=pk)
    try:
        obj.delete()
    except ProtectedError:
        messages.error(
            request, _('Cannot delete "{}" - it is still used by one or more projects.').format(obj.name)
        )
    return redirect(admin_project_status_view)


# Administration - Tags
@permission_required(svjis_edit_admin_tags)
@require_GET
def admin_tag_view(request):
    tag_list = models.Tag.objects.all()
    rows = []
    for t in tag_list:
        rows.append(
            {
                'url': reverse('admin_tag_edit', args=[t.pk]),
                'cells': {'name': t.name},
                'actions': [
                    {'url': reverse('admin_tag_edit', args=[t.pk]), 'icon': 'edit', 'label': _('Edit')},
                    {
                        'url': reverse('admin_tag_delete', args=[t.pk]),
                        'icon': 'delete',
                        'label': _('Delete'),
                        'danger': True,
                        'confirm': f"{_('Do you want to delete')} {t.name} ?",
                    },
                ],
            }
        )

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('tags', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = tag_list
    ctx['columns'] = [{'key': 'name', 'label': _('Name')}]
    ctx['rows'] = rows
    ctx['table_id'] = 'tag-table'
    ctx['desc_id'] = 'tags'
    ctx['list_search'] = True
    return render(request, "admin_tag.html", ctx)


@permission_required(svjis_edit_admin_tags)
@require_GET
def admin_tag_edit_view(request, pk):
    if pk:
        i = get_object_or_404(models.Tag, pk=pk)
        form = forms.TagForm(instance=i)
    else:
        form = forms.TagForm

    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['form'] = form
    ctx['pk'] = pk
    ctx['aside_menu_items'] = get_side_menu('tags', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    return render(request, "admin_tag_edit.html", ctx)


@permission_required(svjis_edit_admin_tags)
@require_POST
def admin_tag_save_view(request):
    pk = int(request.POST['pk'])
    if not pk:
        form = forms.TagForm(request.POST)
    else:
        instance = get_object_or_404(models.Tag, pk=pk)
        form = forms.TagForm(request.POST, instance=instance)

    if not form.is_valid():
        for error in form.errors:
            messages.error(request, f"{_('Form validation error')}: {error}")
        return redirect(admin_tag_view)

    form.save()
    messages.info(request, _('Saved'))
    return redirect(admin_tag_view)


@permission_required(svjis_edit_admin_tags)
@require_GET
def admin_tag_delete_view(request, pk):
    obj = get_object_or_404(models.Tag, pk=pk)
    obj.delete()
    return redirect(admin_tag_view)


# Administration - Waiting messages
@permission_required(svjis_view_admin_menu)
@require_GET
def admin_messages_view(request):
    message_list = models.MessageQueue.objects.filter(status=0)
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('messages', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['object_list'] = message_list.order_by('pk')
    return render(request, "admin_messages.html", ctx)


# Administration - About application
@permission_required(svjis_view_admin_menu)
@require_GET
def admin_about_view(request):
    ctx = utils.get_context()
    ctx['aside_menu_name'] = _("Administration")
    ctx['aside_menu_items'] = get_side_menu('about', request.user)
    ctx['tray_menu_items'] = utils.get_tray_menu('admin', request.user)
    ctx['django_version'] = django_version
    ctx['python_version'] = sys.version
    ctx['svjis_homepage'] = __homepage_url__
    ctx['svjis_repository'] = __repository_url__
    ctx['svjis_issues'] = __issues_url__
    ctx['svjis_translations'] = __translations_url__
    return render(request, "admin_about.html", ctx)
